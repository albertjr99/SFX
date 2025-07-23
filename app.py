from flask import Flask, render_template, request, redirect, session, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from flask_migrate import Migrate
from datetime import datetime, timedelta
import openpyxl
from flask import abort, jsonify
from datetime import datetime, date
import os
from werkzeug.utils import secure_filename
import calendar
import re
from sqlalchemy import func, or_


# --- Configuração da aplicação ---
app = Flask(__name__)
app.secret_key = 'chave-secreta'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///producao.db'

db = SQLAlchemy(app)
migrate = Migrate(app, db)

login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Meses em Português para filtro
MESES_PT = [
    'Janeiro','Fevereiro','Março','Abril','Maio','Junho',
    'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro'
]
class Usuario(db.Model, UserMixin):
    __tablename__ = 'usuario'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    senha = db.Column(db.String(200), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # analista, estagiaria, gerente
    modalidade = db.Column(db.String(20))           # presencial ou teletrabalho
    primeiro_acesso_realizado = db.Column(db.Boolean, default=False)

class LinhaProducao(db.Model):
    __tablename__ = 'linha_producao'
    id                    = db.Column(db.Integer, primary_key=True)
    usuario_id       = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    mes              = db.Column(db.String(20), nullable=False)
    semana           = db.Column(db.String(50), nullable=False)
    indice_linha          = db.Column(db.Integer)
    data_registro         = db.Column(db.DateTime, default=datetime.utcnow)

    numero_processo       = db.Column(db.String(100))
    nome_do_segurado     = db.Column(db.String(100))
    status                = db.Column(db.String(20))
    setor                 = db.Column(db.String(100))
    
    observacoes      = db.Column(db.Text)

    homologacao           = db.Column(db.Boolean, default=False)
    diligencia           = db.Column(db.Boolean, default=False)
    fixacao_civil           = db.Column(db.Boolean, default=False)
    conf_civil            = db.Column(db.Boolean, default=False)
    registro_civil        = db.Column(db.Boolean, default=False)
    conf_civil   = db.Column(db.Boolean, default=False)
    fixacao_pensao           = db.Column(db.Boolean, default=False)
    conf_pensao      = db.Column(db.Boolean, default=False)
    in_68                 = db.Column(db.Boolean, default=False)
    outros                = db.Column(db.Boolean, default=False)
    rg_cpf_doc               = db.Column(db.Boolean, default=False)
    comp_residencia_doc      = db.Column(db.Boolean, default=False)
    cert_nasc_casamento_doc  = db.Column(db.Boolean, default=False)
    requerimento_doc         = db.Column(db.Boolean, default=False)
    dtc_doc                  = db.Column(db.Boolean, default=False)
    negativa_pad_doc         = db.Column(db.Boolean, default=False)
    contracheque_doc         = db.Column(db.Boolean, default=False)
    declaracao_chefia_doc    = db.Column(db.Boolean, default=False)
    termo_subsidio_doc       = db.Column(db.Boolean, default=False)
    declaracao_bens_doc      = db.Column(db.Boolean, default=False)
    extrato_doc              = db.Column(db.Boolean, default=False)
    qd_elevacao_doc          = db.Column(db.Boolean, default=False)
    lotacao_doc              = db.Column(db.Boolean, default=False)
    tempo_contribuicao_doc   = db.Column(db.Boolean, default=False)
    anexos_tce_doc           = db.Column(db.Boolean, default=False)
    portaria_doc             = db.Column(db.Boolean, default=False)
    fixacao_o_doc            = db.Column(db.Boolean, default=False)
    siarhes_doc              = db.Column(db.Boolean, default=False)

  
    # **NOVOS CAMPOS**:
    resolucao          = db.Column(db.Text, default='')
    removido           = db.Column(db.Boolean, default=False)
    consideracoes = db.Column(db.Text, nullable=True)

    usuario = db.relationship('Usuario', backref='linhas')

# Gera as semanas úteis de um mês e ano
def gerar_semanas(mes_num, ano):
    semanas = []
    primeiro_dia = datetime(ano, mes_num, 1)
    ultimo_dia = datetime(ano, mes_num, calendar.monthrange(ano, mes_num)[1])
    semana = []
    dia = primeiro_dia
    while dia <= ultimo_dia:
        if dia.weekday() < 5:
            semana.append(dia.strftime('%d/%m'))
        if dia.weekday() == 4 or dia == ultimo_dia:
            if semana:
                semanas.append(f"{semana[0]} a {semana[-1]}")
                semana = []
        dia += timedelta(days=1)
    return semanas

# Alias para compatibilidade
obter_semanas_do_mes = gerar_semanas




# --- Flask-Login user loader ---
@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# --- Rotas de autenticação ---
@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')
    nome = request.form['nome']
    senha = request.form['senha']
    usuario = Usuario.query.filter(func.lower(Usuario.nome)==nome.lower()).first()
    if usuario and check_password_hash(usuario.senha, senha):
        login_user(usuario)
        session['usuario_id'] = usuario.id
        session['usuario_tipo'] = usuario.tipo
        if usuario.tipo == 'analista':
            return redirect(url_for('registrar_producao'))
        elif usuario.tipo == 'estagiaria':
            return redirect(url_for('painel_estagiarias'))
        else:
            return redirect(url_for('painel_gerente'))
    return 'Login inválido'

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect(url_for('login'))

@app.route('/primeiro-acesso', methods=['POST'])
def primeiro_acesso():
    nome = request.form.get('nome')
    nova_senha = request.form.get('nova_senha')
    # busca usuario pelo nome, ignorando maiúsculas/minúsculas
    usuario = Usuario.query.filter(func.lower(Usuario.nome) == nome.lower()).first()
    if not usuario:
        flash('Usuário não encontrado.')
    elif usuario.primeiro_acesso_realizado:
        flash('Senha já definida anteriormente.')
    else:
        usuario.senha = generate_password_hash(nova_senha)
        usuario.primeiro_acesso_realizado = True
        db.session.commit()
        flash('Senha definida com sucesso. Agora faça login.')
    return redirect(url_for('login'))
    
# ... (seu setup do app, db, Usuario, LinhaProducao, gerar_semanas, etc.) ...



from datetime import datetime, date
from sqlalchemy import or_

@app.route('/pendencias', methods=['GET','POST'])
@login_required
def pendencias():
    if current_user.tipo != 'analista':
        flash('Acesso não autorizado.', 'error')
        return redirect(url_for('login'))

    # Gera lista de meses de Dez/2023 a Dez/2026
    meses = gerar_lista_de_meses(
        inicio=date(2023, 12, 1),
        fim   =date(2026, 12, 1)
    )
    selected_mes = request.args.get('mes', meses[0])

    if request.method == 'POST':
        action  = request.form.get('action')
        proc_id = request.form.get('proc_id', type=int)

        if action == 'add':
            mes  = request.form['mes']
            semana = request.form['semana']
            numero = request.form['numero_processo']
            nome   = request.form['nome_do_segurado']
            status = request.form['status']
            nova = LinhaProducao(
                usuario_id       = current_user.id,
                mes              = mes,
                semana           = semana,
                numero_processo  = numero,
                nome_do_segurado = nome,
                status           = status,
                data_registro    = datetime.utcnow()
            )
            db.session.add(nova)
            db.session.commit()
            flash('Pendência adicionada com sucesso.', 'success')

        else:
            p = LinhaProducao.query.get_or_404(proc_id)
            if p.usuario_id != current_user.id:
                flash('Operação não permitida.', 'error')
                return redirect(url_for('pendencias', mes=selected_mes))

            if action == 'toggle_removido':
                p.removido = not p.removido
                db.session.commit()
                flash('Pendência atualizada.', 'success')

            elif action == 'save_consideracoes':
                p.consideracoes = request.form.get('consideracoes','').strip()
                db.session.commit()
                flash('Considerações salvas com sucesso.', 'success')

            elif action == 'save_resolucao':
                p.resolucao = request.form.get('resolucao','').strip()
                db.session.commit()
                flash('Resolução salva com sucesso.', 'success')

        return redirect(url_for('pendencias', mes=selected_mes))

    # GET: listar pendências
    pend_status = ['tramitado', 'aguardando', 'acompanhar']
    procs = (
        LinhaProducao.query
          .filter(LinhaProducao.usuario_id == current_user.id)
          .filter(
              or_(
                  LinhaProducao.status.in_(pend_status),
                  LinhaProducao.removido == True
              )
          )
          .order_by(LinhaProducao.data_registro.asc())
          .all()
    )

    labels = {
      'tramitado': 'Tramitado – aguardando',
      'aguardando': 'Aguardando orientação Chefe',
      'acompanhar': 'Acompanhar mesmo após envio'
    }

    return render_template('pendencias.html',
                           procs=procs,
                           labels=labels,
                           meses=meses,
                           selected_mes=selected_mes
    )


@app.route('/buscar_producao')
@login_required
def buscar_producao():
    q = request.args.get('q', '').strip()
    resultados = []
    if q:
        # meses fixos do sistema
        meses = ['Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
        ano = datetime.now().year
        for mes in meses:
            # gera as semanas desse mês
            idx = MESES_PT.index(mes) + 1
            semanas = gerar_semanas(idx, ano)
            for semana in semanas:
                # busca todos os registros que coincidem com número ou nome
                prods = (LinhaProducao.query
                         .filter_by(mes=mes, semana=semana)
                         .filter(
                           (LinhaProducao.numero_processo.ilike(f'%{q}%')) |
                           (LinhaProducao.nome_do_segurado.ilike(f'%{q}%'))
                         )
                         .all())
                for p in prods:
                    resultados.append({
                        'mes': mes,
                        'semana': semana,
                        'numero_processo': p.numero_processo or '',
                        'nome_do_segurado': p.nome_do_segurado or ''
                    })
    return jsonify(resultados=resultados)

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.route('/acompanhamento-anual')
@login_required
def acompanhamento_anual():
    campos = [ 'homologacao', 'fixacao_civil', 'registro_civil',
        'conf_civil', 'conf_pensao', 'fixacao_pensao', 'in_68', 'outros', 'diligencia' ]
    meses = ['Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    totais_anuais = { m: {c:0 for c in campos} for m in meses }
    for p in LinhaProducao.query.filter(LinhaProducao.mes.in_(meses)).all():
        for c in campos:
            if getattr(p, c):
                totais_anuais[p.mes][c] += 1

    grafico_anos = { c: sum(m[c] for m in totais_anuais.values()) for c in campos }

    return render_template(
        'acompanhamento_anual.html',
        totais_anuais = totais_anuais,
        grafico_anos   = grafico_anos,
        meses          = meses,
        campos         = campos
    )

@app.route('/acompanhamento-analista')
@login_required
def acompanhamento_analista():
    # só gestores
    if current_user.tipo == 'analista':
        flash('Acesso não autorizado.')
        return redirect(url_for('index'))

    analista_id      = request.args.get('analista_id', type=int)
    # lista fixa de meses (meses que você usa no sistema)
    meses            = ['Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    # define mês padrão: parâmetro ou mês corrente (se estiver na lista) ou 'Junho'
    hoje_nome        = datetime.now().strftime('%B').capitalize()
    mes_param        = request.args.get('mes')
    if mes_param in meses:
        mes = mes_param
    elif hoje_nome in meses:
        mes = hoje_nome
    else:
        mes = 'Junho'

    # gerar semanas do mês
    ano              = datetime.now().year
    mes_idx          = meses.index(mes)
    semanas          = gerar_semanas(mes_idx + 6, ano)

    # parâmetros de filtro do gráfico
    view             = request.args.get('view', 'mes')   # 'semana','mes','ano'
    selected_semana  = request.args.get('semana', semanas[0] if semanas else '')

    # campos de contagem
    campos = [
        'homologacao', 'fixacao_civil', 'registro_civil',
        'conf_civil', 'conf_pensao', 'fixacao_pensao', 'in_68', 'outros', 'diligencia'
    ]

    # busca o analista
    analista = Usuario.query.get(analista_id) if analista_id else None

    # inicializa contadores
    counts = {c: 0 for c in campos}

    if analista:
        if view == 'semana':
            # só aquela semana
            for p in LinhaProducao.query.filter_by(
                    usuario_id=analista.id, mes=mes, semana=selected_semana):
                for c in campos:
                    if getattr(p, c):
                        counts[c] += 1

        elif view == 'mes':
            # soma todas as semanas do mês
            for s in semanas:
                for p in LinhaProducao.query.filter_by(
                        usuario_id=analista.id, mes=mes, semana=s):
                    for c in campos:
                        if getattr(p, c):
                            counts[c] += 1

        else:  # view == 'ano'
            # soma todos os meses
            for m in meses:
                for p in LinhaProducao.query.filter_by(
                        usuario_id=analista.id, mes=m):
                    for c in campos:
                        if getattr(p, c):
                            counts[c] += 1

    return render_template(
        'acompanhamento_analista.html',
        analista=analista,
        meses=meses,
        mes=mes,
        semanas=semanas,
        view=view,
        selected_semana=selected_semana,
        campos=campos,
        counts=counts
    )

from datetime import datetime

# no topo do arquivo, depois de `app = Flask(__name__)`
@app.template_filter('getattr')
def jinja2_getattr(obj, name):
    """Permite em templates usar obj|getattr('campo') para chamar getattr(obj,'campo')."""
    return getattr(obj, name,)


@app.route('/painel-gerente', methods=['GET', 'POST'])
@login_required
def painel_gerente():
    # 1) só gestores acessam
    if current_user.tipo == 'analista':
        flash('Acesso não autorizado.')
        return redirect(url_for('index'))

    # 2) listas de analistas
    analistas_presencial   = Usuario.query.filter_by(tipo='analista', modalidade='presencial').all()
    analistas_teletrabalho = Usuario.query.filter_by(tipo='analista', modalidade='teletrabalho').all()

    # 3) leitura dos filtros
    analista_id      = request.values.get('analista_id', type=int)
    mes_param        = request.values.get('mes')
    selected_semana  = request.values.get('semana', 'Mês inteiro')

    # meses que trabalhamos
    meses = ['Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    # define mês
    if mes_param in meses:
        mes = mes_param
    else:
        m_atual = datetime.now().month
        mes = meses[m_atual-6] if 6 <= m_atual <= 12 else meses[0]

    usuario_sel = Usuario.query.get(analista_id) if analista_id else None

    # gera semanas e linhas
    ano     = datetime.now().year
    idx     = meses.index(mes)
    semanas = gerar_semanas(idx + 6, ano)
    linhas  = 15  # tanto presencial quanto teletrabalho

    campos = [
        'homologacao', 'fixacao_civil', 'registro_civil',
        'conf_civil', 'conf_pensao', 'fixacao_pensao', 'in_68', 'outros', 'diligencia'
    ]

    # init
    processos_info  = {}
    totais_semanais = {s: {c:0 for c in campos} for s in semanas}
    total_feito     = 0
    meta            = 100
    percentual_meta = 0
    alertas         = {}

    if usuario_sel:
        # meta mensal
        meta = 112 if usuario_sel.modalidade=='teletrabalho' else 100

        # --- POST: salva todas as edições ---
        if request.method == 'POST':
            for s in semanas:
                for i in range(linhas):
                    # busca ou cria linha
                    p = (LinhaProducao.query
                         .filter_by(usuario_id=usuario_sel.id, mes=mes, semana=s)
                         .offset(i).first())
                    if not p:
                        p = LinhaProducao(
                            usuario_id    = usuario_sel.id,
                            mes           = mes,
                            semana        = s,
                            indice_linha  = i,
                            data_registro = datetime.utcnow()
                        )
                        db.session.add(p)

                    # campos de texto
                    p.numero_processo = request.form.get(f'{s}_{i}_numero_processo','')
                    p.nome_do_segurado      = request.form.get(f'{s}_{i}_nome_do_segurado','')
                    p.setor = request.form.get(f'{s}_{i}_setor', '')
                    p.observacoes    = request.form.get(f'{s}_{i}_observacoes','')

                    # checkboxes
                    for c in campos:
                        setattr(p, c,
                                bool(request.form.get(f'{s}_{i}_{c}')))

            db.session.commit()
            flash('Produção atualizada com sucesso.')
            return redirect(url_for('painel_gerente',
                                    analista_id=usuario_sel.id,
                                    mes=mes,
                                    semana=selected_semana))

        # monta processos_info e totais
        processos_info = {
            s: [
                (LinhaProducao.query
                    .filter_by(usuario_id=usuario_sel.id, mes=mes, semana=s)
                    .offset(i).first())
                for i in range(linhas)
            ]
            for s in semanas
        }
        for s in semanas:
            for p in LinhaProducao.query.filter_by(usuario_id=usuario_sel.id, mes=mes, semana=s):
                for c in campos:
                    if getattr(p, c):
                        totais_semanais[s][c] += 1

        # total e progresso
        total_feito     = sum(sum(vals.values()) for vals in totais_semanais.values())
        percentual_meta = min(int((total_feito / meta) * 100), 100)

        # alertas
        for s in semanas:
            feito    = sum(totais_semanais[s].values())
            esperado = 15
            if feito < esperado:
                alertas[s] = f"Faltam {esperado - feito} tarefas"

    # totais anuais p/ comparativo
    totais_anuais = {m: {c:0 for c in campos} for m in meses}
    for m in meses:
        for p in LinhaProducao.query.filter_by(mes=m):
            for c in campos:
                if getattr(p, c):
                    totais_anuais[m][c] += 1

    # mini-relatório mensal
    total_mes = {c: sum(totaiss[c] for totaiss in totais_semanais.values())
                 for c in campos}

    return render_template(
        'painel_gerente.html',
        analistas_presencial   = analistas_presencial,
        analistas_teletrabalho = analistas_teletrabalho,
        usuario_selecionado    = usuario_sel,
        meses                  = meses,
        mes                    = mes,
        semanas                = semanas,
        linhas                 = linhas,
        campos                 = campos,
        processos_info         = processos_info,
        totais                 = totais_semanais,
        total_mes              = total_mes,
        total_feito            = total_feito,
        meta                   = meta,
        percentual_meta        = percentual_meta,
        selected_semana        = selected_semana,
        alertas                = alertas,
        totais_anuais          = totais_anuais
    )
# --- Acompanhamento Pessoal com filtro de mês e semana ---
@app.route('/acompanhamento-pessoal')
@login_required
def acompanhamento_pessoal():
    # acesso apenas para analistas
    if current_user.tipo != 'analista':
        flash('Acesso não autorizado.')
        return redirect(url_for('login'))

    # lista de meses para filtro
    meses = MESES_PT

    # parâmetros GET opcionais
    selected_mes = request.args.get('mes') or MESES_PT[datetime.now().month - 1]
    selected_semana = request.args.get('semana') or 'Mês inteiro'

    # converte mes para número
    mes_num = MESES_PT.index(selected_mes) + 1
    ano = datetime.now().year

    # gera semanas do mês selecionado
    semanas = gerar_semanas(mes_num, ano)

    campos = ['homologacao', 'fixacao_civil', 'registro_civil',
        'conf_civil', 'conf_pensao', 'fixacao_pensao', 'in_68', 'outros', 'diligencia']

    totais = {}
    total_mes = {c: 0 for c in campos}
    total_feito = 0

    # conta por semana e total do mês
    for s in semanas:
        cont = {c: 0 for c in campos}
        prods = LinhaProducao.query.filter_by(
            usuario_id=current_user.id,
            mes=selected_mes,
            semana=s
        ).all()
        for p in prods:
            for c in campos:
                if getattr(p, c, False):
                    cont[c] += 1
                    total_mes[c] += 1
                    total_feito += 1
        totais[s] = cont

    meta = 112 if current_user.modalidade == 'teletrabalho' else 100
    percentual_meta = round((total_feito / meta) * 100, 1) if meta else 0

    return render_template('acompanhamento_pessoal.html',
        usuario=current_user,
        meses=meses,
        semanas=semanas,
        selected_mes=selected_mes,
        selected_semana=selected_semana,
        campos=campos,
        totais=totais,
        total_mes=total_mes,
        total_feito=total_feito,
        meta=meta,
        percentual_meta=percentual_meta
    )

@app.route('/editar-producao/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_producao(id):
    producao = LinhaProducao.query.get_or_404(id)
    if request.method == 'POST':
        producao.numero_processo = request.form['numero_processo']
        producao.nome_do_segurado = request.form['nome_do_segurado']
        producao.setor = request.form['setor']
        producao.observacoes = request.form['observacoes']
        db.session.commit()
        return redirect(url_for('painel_gerente'))
    return render_template('editar_producao.html', producao=producao)

# Continuação será feita na próxima etapa

# ================================
# Registrar Produção (painel do analista)
# ================================
# Registrar Produção (painel do analista)
# ================================
@app.route('/registrar-producao', methods=['GET', 'POST'])
@login_required
def registrar_producao():
    if current_user.tipo != 'analista':
        flash('Acesso não autorizado.')
        return redirect(url_for('login'))

    # --- Configuração inicial ---
    usuario = Usuario.query.get(current_user.id)
    meses   = ['Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    ano     = datetime.now().year

    # escolhe mês
    mes_param = request.args.get('mes')
    if mes_param in meses:
        mes_atual = mes_param
    else:
        m = datetime.now().month
        mes_atual = meses[m-6] if 6 <= m <= 12 else meses[0]

    idx     = meses.index(mes_atual)
    semanas = gerar_semanas(idx + 6, ano)

    default_linhas = 15

    # campos de produção e de conferência documental
    campos_cb  = [
        'homologacao','fixacao_civil','in_68','conf_civil',
        'registro_civil','fixacao_pensao','conf_pensao',
        'diligencia','outros'
    ]
    campos_doc = [
        'rg_cpf_doc','comp_residencia_doc','cert_nasc_casamento_doc',
        'requerimento_doc','dtc_doc','negativa_pad_doc',
        'contracheque_doc','declaracao_chefia_doc','termo_subsidio_doc',
        'declaracao_bens_doc','extrato_doc','qd_elevacao_doc',
        'lotacao_doc','tempo_contribuicao_doc','anexos_tce_doc',
        'portaria_doc','fixacao_o_doc','siarhes_doc'
    ]

    # --- POST: salvar produção e conferência documental ---
    if request.method == 'POST':
        # 1) monta linhas_por_semana a partir do form
        linhas_por_semana = {}
        for semana in semanas:
            valor = request.form.get(f'{semana}_rows', type=int)
            linhas_por_semana[semana] = max(default_linhas, valor or default_linhas)

        # 2) exclui do DB os registros que não vieram no form
        for semana in semanas:
            enviados = {
                int(request.form.get(f'{semana}_{i}_id'))
                for i in range(linhas_por_semana[semana])
                if request.form.get(f'{semana}_{i}_id')
            }
            todos = LinhaProducao.query.filter_by(
                usuario_id=usuario.id,
                mes=mes_atual,
                semana=semana
            ).all()
            for p in todos:
                if p.id not in enviados:
                    db.session.delete(p)

        # 3) cria ou atualiza cada linha
        for semana in semanas:
            for i in range(linhas_por_semana[semana]):
                p_id = request.form.get(f'{semana}_{i}_id')
                if p_id:
                    p = LinhaProducao.query.get(int(p_id))
                else:
                    p = LinhaProducao(
                        usuario_id   = usuario.id,
                        mes          = mes_atual,
                        semana       = semana,
                        indice_linha = i,
                        data_registro= datetime.utcnow()
                    )
                    db.session.add(p)

                # campos texto / select
                p.status           = request.form.get(f'{semana}_{i}_status','')
                p.numero_processo  = request.form.get(f'{semana}_{i}_numero_processo','')
                p.nome_do_segurado = request.form.get(f'{semana}_{i}_nome_do_segurado','')
                p.setor            = request.form.get(f'{semana}_{i}_setor','')
                p.observacoes      = request.form.get(f'{semana}_{i}_obs','')

                # checkboxes de produção
                for cb in campos_cb:
                    setattr(p, cb, bool(request.form.get(f'{semana}_{i}_{cb}')))

                # checkboxes de conferência documental
                for doc in campos_doc:
                    setattr(p, doc, bool(request.form.get(f'{semana}_{i}_{doc}')))

                # garante ordem persistente
                p.indice_linha = i

        db.session.commit()
        flash('Produção salva com sucesso!')
        return redirect(url_for('registrar_producao', mes=mes_atual))

    # --- GET: monta linhas_por_semana a partir do banco ---
    linhas_por_semana = {}
    for semana in semanas:
        cnt = (LinhaProducao.query
               .filter_by(usuario_id=usuario.id, mes=mes_atual, semana=semana)
               .count())
        linhas_por_semana[semana] = max(default_linhas, cnt)

    # monta processos_info para render
    processos_info = {
        semana: [
            (LinhaProducao.query
             .filter_by(usuario_id=usuario.id, mes=mes_atual, semana=semana)
             .order_by(LinhaProducao.indice_linha)
             .offset(i).first())
            for i in range(linhas_por_semana[semana])
        ]
        for semana in semanas
    }

    # cálculo de totais (inalterado)
    totais      = {s: {cb:0 for cb in campos_cb} for s in semanas}
    total_feito = 0
    for s in semanas:
        for p in processos_info[s] or []:
            if not p: continue
            for cb in campos_cb:
                if getattr(p, cb):
                    totais[s][cb] += 1
                    total_feito += 1

    meta            = 112 if usuario.modalidade=='teletrabalho' else 100
    percentual_meta = min(int((total_feito / meta) * 100), 100) if meta else 0

    return render_template(
        'registrar_producao.html',
        usuario           = usuario,
        semanas           = semanas,
        linhas_por_semana = linhas_por_semana,
        processos_info    = processos_info,
        campos            = campos_cb,
        campos_doc        = campos_doc,
        totais            = totais,
        total_feito       = total_feito,
        meta              = meta,
        percentual_meta   = percentual_meta,
        mes_anterior      = meses[idx-1] if idx>0 else meses[-1],
        mes_atual         = mes_atual,
        mes_posterior     = meses[idx+1] if idx<len(meses)-1 else meses[0],
    )



@app.route('/plano-acao', methods=['GET','POST'])
@login_required
def plano_acao():
    # só o Lucas pode acessar
    if current_user.nome.lower() != 'lucas':
        abort(403)

    # lê o parâmetro ?ano= (default 2024)
    ano_req = request.args.get('ano', '2024')
    if ano_req not in ('2024', '2025'):
        ano_req = '2024'

    # caminho da planilha
    excel_path = os.path.join(app.root_path, 'data', 'Pasta2.xlsx')

    # salvamento via POST
    if request.method == 'POST':
        payload = request.get_json() or {}
        ano_post = payload.get('ano', ano_req)
        new_data = payload.get('data', [])
        # reabre para escrita
        wb2 = openpyxl.load_workbook(excel_path)
        ws2 = wb2['Controle processos 2025']
        # define intervalo de linhas conforme ano
        if ano_post == '2024':
            mr, Mr = 4, 15
        else:
            mr, Mr = 16, 28
        # escreve cada valor
        for i, row_vals in enumerate(new_data):
            for j, val in enumerate(row_vals):
                ws2.cell(row=mr + i, column=j + 1).value = val
        wb2.save(excel_path)
        return jsonify(success=True)

    # GET: carrega a planilha
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb['Controle processos 2025']

    # cabeçalhos: linha 2 (colunas A–I)
    headers = [cell.value for cell in ws[2][:9]]

    # define intervalo de linhas conforme ano
    if ano_req == '2024':
        min_row, max_row = 4, 15
        info_cell = 'A31'
    else:
        min_row, max_row = 16, 28
        info_cell = 'A31'

    # dados do mês selecionado com formatação de data
    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=9, values_only=True):
        row = list(row)
        raw = row[7]
        formatted = ''
        if isinstance(raw, (datetime, date)):
            formatted = raw.strftime('%d/%m/%Y')
        elif isinstance(raw, str) and raw.strip():
            try:
                d = datetime.fromisoformat(raw.split()[0])
                formatted = d.strftime('%d/%m/%Y')
            except:
                formatted = raw
        row[7] = formatted
        data.append(row)

    # texto laranja de contexto
    info_text = ws[info_cell].value

    # tabela de chamados permanece inalterada
    header_row_chamados = 18
    chamado_headers = [cell.value for cell in ws[header_row_chamados][:7]]
    chamado_data = [
        list(r)
        for r in ws.iter_rows(
            min_row=19,
            max_row=28,
            max_col=7,
            values_only=True
        )
        if any(c != "" for c in r)
    ]

    return render_template(
        'plano_acao.html',
        headers=headers,
        data=data,
        info_text=info_text,
        chamado_headers=chamado_headers,
        chamado_data=chamado_data,
        ano=ano_req
    )

@app.template_filter('get_attr')
def get_attr(obj, name):
    return getattr(obj, name, None)


# ================================
# Painel das Estagiárias (edição em lote)
# ================================
@app.route('/painel-estagiarias', methods=['GET', 'POST'])
@login_required
def painel_estagiarias():
    if current_user.tipo != 'estagiaria':
        flash('Acesso não autorizado.')
        return redirect(url_for('index'))

    analistas_presencial   = Usuario.query.filter_by(tipo='analista', modalidade='presencial').all()
    analistas_teletrabalho = Usuario.query.filter_by(tipo='analista', modalidade='teletrabalho').all()

    # recebe tanto GET quanto POST
    analista_id = request.values.get('analista_id', type=int)
    mes         = request.values.get('mes') or 'Junho'

    usuario_sel = Usuario.query.get(analista_id) if analista_id else None

    ano    = 2025
    meses  = ['Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    idx    = meses.index(mes)
    semanas = gerar_semanas(idx + 6, ano)
    linhas  = 15

    campos_texto    = ['numero_processo','nome_do_segurado','setor', 'observacoes']
    campos_checkbox = [
        'homologacao', 'fixacao_civil', 'registro_civil',
        'conf_civil', 'conf_pensao', 'fixacao_pensao', 'in_68', 'outros', 'diligencia'
    ]

    # POST: salva lotes
    if request.method == 'POST' and usuario_sel:
        for s in semanas:
            for i in range(linhas):
                p = (LinhaProducao.query
                     .filter_by(usuario_id=usuario_sel.id, mes=mes, semana=s)
                     .offset(i).first())
                if not p:
                    p = LinhaProducao(
                        usuario_id    = usuario_sel.id,
                        mes           = mes,
                        semana        = s,
                        indice_linha  = i,
                        data_registro = datetime.utcnow()
                    )
                    db.session.add(p)

                # texto
                p.numero_processo = request.form.get(f'{s}_{i}_numero_processo','')
                p.nome_do_segurado      = request.form.get(f'{s}_{i}_nome_do_segurado','')
                p.setor           = request.form.get(f'{s}_{i}_','')
                p.observacoes     = request.form.get(f'{s}_{i}_observacoes','')

                # checkboxes
                for c in campos_checkbox:
                    setattr(p, c,
                            bool(request.form.get(f'{s}_{i}_{c}')))

        db.session.commit()
        flash('Edições salvas com sucesso!')
        return redirect(url_for('painel_estagiarias',
                                analista_id=usuario_sel.id,
                                mes=mes))

    # GET: carrega info atual
    processos_info = {}
    if usuario_sel:
        processos_info = {
            s: [
                (LinhaProducao.query
                 .filter_by(usuario_id=usuario_sel.id, mes=mes, semana=s)
                 .offset(i).first())
                for i in range(linhas)
            ]
            for s in semanas
        }

    return render_template(
        'painel_estagiarias.html',
        analistas_presencial   = analistas_presencial,
        analistas_teletrabalho = analistas_teletrabalho,
        usuario_selecionado    = usuario_sel,
        mes                    = mes,
        semanas                = semanas,
        linhas                 = linhas,
        campos                 = campos_texto,
        processos_info         = processos_info
    )


@app.route('/relatorio-geral')
@login_required
def relatorio_geral():
    if 'usuario_id' not in session or session['usuario_tipo'] == 'analista':
        flash('Acesso não autorizado.')
        return redirect(url_for('index'))

    ano = 2025
    meses = ['Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    campos = [
        'homologacao', 'fixacao_civil', 'registro_civil',
        'conf_civil', 'conf_pensao', 'fixacao_pensao', 'in_68', 'outros', 'diligencia'
    ]

    totais_gerais = {mes: {campo: 0 for campo in campos} for mes in meses}

    for mes in meses:
        producoes = LinhaProducao.query.filter_by(mes=mes).all()
        for producao in producoes:
            for campo in campos:
                if getattr(producao, campo):
                    totais_gerais[mes][campo] += 1

    return render_template(
        'relatorio_geral.html',
        totais_gerais=totais_gerais,
        campos=campos,
        meses=meses
    )

# pasta onde os PDFs serão salvos
UPLOAD_FOLDER = os.path.join(app.root_path, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/conferencia-documental/<int:processo_id>', methods=['GET', 'POST'])
@login_required
def conferencia_documental(processo_id):
    # Carrega o processo ou 404 se não existir
    p = LinhaProducao.query.get_or_404(processo_id)

    # Defina aqui a lista de documentos que você quer checar
    documentos = [
        'RG', 'CPF', 'Comprovante de Endereço',
        'Procuração', 'Contrato', # etc...
    ]

    if request.method == 'POST':
        # Para cada documento, salva sim/não no atributo do modelo
        for doc in documentos:
            status = request.form.get(f'status_{doc}', 'não')
            setattr(p, f'{doc.lower().replace(" ", "_")}_doc', (status == 'sim'))

        # Se vier PDF anexado, salva no disco e guarda o nome no atributo `pdf_anexo`
        file = request.files.get('pdf_anexo')
        if file and file.filename:
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            p.pdf_anexo = filename

        db.session.commit()
        flash('Conferência documental atualizada com sucesso!', 'success')
        return redirect(url_for('registrar_producao'))

    return render_template(
        'conferencia_documental.html',
        processo=p,
        documentos=documentos
    )

@app.route('/editar-producao-lote/<int:analista_id>', methods=['POST'])
@login_required
def editar_producao_lote(analista_id):
    if 'usuario_id' not in session or session['usuario_tipo'] != 'estagiaria':
        flash('Acesso não autorizado.')
        return redirect(url_for('index'))

    usuario = Usuario.query.get_or_404(analista_id)
    mes = request.args.get('mes') or 'Junho'
    ano = 2025
    meses = ['Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    mes_index = meses.index(mes)
    semanas = gerar_semanas(mes_index + 6, ano)
    linhas = 15 if usuario.modalidade == 'teletrabalho' else 15

    for semana in semanas:
        for i in range(linhas):
            processo = LinhaProducao.query.filter_by(
                usuario_id=analista_id, semana=semana, mes=mes
            ).offset(i).first()

            if not processo:
                processo = LinhaProducao(
                    usuario_id=analista_id,
                    mes=mes,
                    semana=semana,
                    data_registro=datetime.utcnow()
                )
                db.session.add(processo)

            processo.numero_processo = request.form.get(f'{semana}_{i}_numero_processo') or ""
            processo.nome_do_segurado = request.form.get(f'{semana}_{i}_nome_do_segurado') or ""
            processo.setor = request.form.get(f'{semana}_{i}_') or ""
            processo.observacoes = request.form.get(f'{s}_{i}_observacoes', '')

    db.session.commit()
    flash('Edições salvas com sucesso.')
    return redirect(url_for('painel_estagiarias', analista_id=analista_id, mes=mes))

if __name__ == '__main__':
    app.run(debug=True)
